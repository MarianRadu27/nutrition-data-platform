#!/usr/bin/env python3
"""Import ANSES/Ciqual foods and nutrient values.

The ANSES workbook stores one food per row and nutrient values in many
columns. This script imports the food rows into source_foods and the nutrient
matrix into source_food_nutrient_values.

Use --dry-run first to preview counts without writing to the database.
"""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
import os
from pathlib import Path
import re
import sys
from typing import Any

import pymysql
from dotenv import load_dotenv
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[3]
BACKEND_DIR = PROJECT_ROOT / "backend"

ANSES_FILE = (
    PROJECT_ROOT
    / "temp"
    / "EuropeNutrientsDBs"
    / "anses"
    / "Table Ciqual 2025_ENG_2025_11_03.xlsx"
)

ANSES_DATA_SOURCE_CODE = "ANSES_CIQUAL"
ANSES_REFERENCE_CODE = "ANSES_CIQUAL_2025_WORKBOOK"
ANSES_REFERENCE_TEXT = (
    "ANSES/Ciqual 2025 food composition workbook: "
    "Table Ciqual 2025_ENG_2025_11_03.xlsx"
)

FOOD_COMPOSITION_SHEET_NAME = "food composition"
INFOODS_SHEET_NAME = "INFOODS codes"

FIRST_NUTRIENT_COLUMN = "Energy,\nRegulation\nEU No\n1169\n2011 (kJ\n100g)"
JONES_FACTOR_COLUMN = "Jones\nfactor"

FOOD_CODE_COLUMN = "alim_code"
FOOD_NAME_EN_COLUMN = "alim_nom_eng"
SCIENTIFIC_NAME_COLUMN = "alim_nom_sci"
CATEGORY_EN_COLUMN = "alim_grp_nom_eng"
SUBCATEGORY_EN_COLUMN = "alim_ssgrp_nom_eng"
SUBSUBCATEGORY_EN_COLUMN = "alim_ssssgrp_nom_eng"

INFOODS_TAG_COLUMN = "INFDSTAG"
INFOODS_SOURCE_CODE_COLUMN = "ORIGCPCD"
INFOODS_NAME_COLUMN = "const_nom_eng"

REQUIRED_FOOD_COLUMNS = [
    FOOD_CODE_COLUMN,
    FOOD_NAME_EN_COLUMN,
    SCIENTIFIC_NAME_COLUMN,
    CATEGORY_EN_COLUMN,
    SUBCATEGORY_EN_COLUMN,
    SUBSUBCATEGORY_EN_COLUMN,
]

REQUIRED_INFOODS_COLUMNS = [
    INFOODS_TAG_COLUMN,
    INFOODS_SOURCE_CODE_COLUMN,
    INFOODS_NAME_COLUMN,
]

INSERT_CHUNK_SIZE = 5000
BELOW_LIMIT_CALC_VALUE = Decimal("0")


@dataclass(frozen=True)
class SourceFood:
    """One ANSES food definition."""

    source_food_code: str
    food_name_original: str | None
    food_name_en: str | None
    food_name_ro: str | None
    category_original: str | None
    category_en: str | None
    category_ro: str | None
    basis: str | None
    notes: str | None


@dataclass(frozen=True)
class NutrientColumn:
    """One nutrient column mapped to an ANSES source nutrient code."""

    header: str
    source_nutrient_code: str
    unit: str | None


@dataclass(frozen=True)
class AnsesNutrientValue:
    """One ANSES nutrient value after wide-to-long conversion."""

    source_food_code: str
    source_nutrient_code: str
    raw_value: str
    value: Decimal | None
    value_qualifier: str | None
    unit: str | None
    basis: str


def env_str(name: str, default: str) -> str:
    value = os.getenv(name)
    if value is None:
        return default
    stripped = value.strip()
    return stripped if stripped else default


def env_int(name: str, default: int) -> int:
    raw = os.getenv(name)
    if raw is None:
        return default
    try:
        return int(raw.strip())
    except ValueError:
        return default


def get_connection() -> pymysql.connections.Connection:
    """Open a MySQL connection using the backend .env settings."""
    load_dotenv(BACKEND_DIR / ".env")
    return pymysql.connect(
        host=env_str("DB_HOST", "127.0.0.1"),
        port=env_int("DB_PORT", 3307),
        user=env_str("DB_USER", "nutrition"),
        password=env_str("DB_PASSWORD", "nutritionpass"),
        database=env_str("DB_NAME", "nutrition"),
        charset="utf8mb4",
        autocommit=False,
        cursorclass=pymysql.cursors.DictCursor,
    )


def clean_text(value: Any, *, dash_is_missing: bool = False) -> str | None:
    """Return normalized text, treating empty values as missing."""
    if value is None:
        return None

    cleaned = " ".join(str(value).split())
    if not cleaned:
        return None

    if dash_is_missing and cleaned == "-":
        return None

    return cleaned


def normalize_unit(value: str | None) -> str | None:
    """Normalize unit spelling while keeping storage simple."""
    cleaned = clean_text(value)
    if cleaned is None:
        return None

    cleaned = cleaned.replace("\u00c2\u00b5", "u")
    cleaned = cleaned.replace("\u00b5", "u")
    cleaned = cleaned.replace("\u03bc", "u")
    cleaned = cleaned.replace("\ufffd", "u")
    cleaned = cleaned.replace("ï¿½", "u")
    return cleaned.lower()


def parse_anses_unit(source_name: str) -> str | None:
    """Extract the reported unit from an ANSES label such as '(g/100g)'."""
    match = re.search(r"\(([^()]*)\)\s*$", source_name)
    if not match:
        return None

    raw_unit = match.group(1).split("/", 1)[0]
    return normalize_unit(raw_unit)


def normalize_nutrient_name(value: Any) -> str:
    """Normalize ANSES nutrient labels for matching workbook sheets."""
    if value is None:
        return ""

    normalized = str(value).replace("\n", " ")
    normalized = re.sub(r"\s*/\s*", " ", normalized)
    normalized = re.sub(r"\(\s*", "(", normalized)
    normalized = re.sub(r"\s*\)", ")", normalized)
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip().lower()


def validate_columns(columns: list[str], required_columns: list[str]) -> None:
    missing = [column for column in required_columns if column not in columns]
    if missing:
        raise RuntimeError("Missing required columns: " + ", ".join(missing))


def get_columns(worksheet) -> list[str]:
    """Return the first row from a worksheet as strings."""
    return [
        "" if value is None else str(value)
        for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]


def split_metadata_and_nutrient_columns(
    columns: list[str],
) -> tuple[list[str], list[str]]:
    """Split ANSES main columns into food metadata and nutrient columns."""
    first_nutrient_column_index = columns.index(FIRST_NUTRIENT_COLUMN)
    return (
        columns[:first_nutrient_column_index],
        columns[first_nutrient_column_index:],
    )


def read_nutrient_column_map(workbook) -> tuple[list[NutrientColumn], list[str]]:
    """Map ANSES food-composition nutrient columns to INFOODS rows."""
    food_worksheet = workbook[FOOD_COMPOSITION_SHEET_NAME]
    infoods_worksheet = workbook[INFOODS_SHEET_NAME]

    food_columns = get_columns(food_worksheet)
    _, food_nutrient_columns = split_metadata_and_nutrient_columns(food_columns)

    infoods_rows = infoods_worksheet.iter_rows(values_only=True)
    infoods_headers = [
        "" if value is None else str(value)
        for value in next(infoods_rows)
    ]
    validate_columns(infoods_headers, REQUIRED_INFOODS_COLUMNS)

    code_index = infoods_headers.index(INFOODS_SOURCE_CODE_COLUMN)
    name_index = infoods_headers.index(INFOODS_NAME_COLUMN)

    nutrient_by_name: dict[str, tuple[str, str | None]] = {}
    for row in infoods_rows:
        source_code = clean_text(row[code_index])
        source_name = clean_text(row[name_index])
        if source_code is None or source_name is None:
            continue

        nutrient_by_name[normalize_nutrient_name(source_name)] = (
            source_code,
            parse_anses_unit(source_name),
        )

    mapped_columns: list[NutrientColumn] = []
    skipped_columns: list[str] = []

    for header in food_nutrient_columns:
        normalized_header = normalize_nutrient_name(header)
        match = nutrient_by_name.get(normalized_header)
        if match is None:
            skipped_columns.append(header)
            continue

        source_code, unit = match
        mapped_columns.append(
            NutrientColumn(
                header=header,
                source_nutrient_code=source_code,
                unit=unit,
            )
        )

    unexpected_skipped_columns = [
        column for column in skipped_columns if column != JONES_FACTOR_COLUMN
    ]
    if unexpected_skipped_columns:
        raise RuntimeError(
            "Unexpected ANSES nutrient columns without INFOODS mapping: "
            + ", ".join(clean_text(column) or column for column in unexpected_skipped_columns)
        )

    return mapped_columns, skipped_columns


def build_food_notes(row: dict[str, Any]) -> str | None:
    """Keep ANSES category levels that do not fit dedicated source_food columns."""
    note_parts = []
    subcategory = clean_text(row.get(SUBCATEGORY_EN_COLUMN), dash_is_missing=True)
    subsubcategory = clean_text(row.get(SUBSUBCATEGORY_EN_COLUMN), dash_is_missing=True)
    scientific_name = clean_text(row.get(SCIENTIFIC_NAME_COLUMN), dash_is_missing=True)

    if subcategory:
        note_parts.append(f"subcategory_en: {subcategory}")

    if subsubcategory:
        note_parts.append(f"subsubcategory_en: {subsubcategory}")

    if scientific_name:
        note_parts.append(f"scientific_name: {scientific_name}")

    return "\n".join(note_parts) if note_parts else None


def parse_nutrient_value(value: Any) -> tuple[str, Decimal | None, str | None] | None:
    """Convert one ANSES source value while preserving source markers."""
    if value is None:
        return None

    raw_value = clean_text(value)
    if raw_value is None or raw_value == "-":
        return None

    normalized = raw_value.lower()
    if normalized in {"trace", "traces", "tr", "t"}:
        return raw_value, BELOW_LIMIT_CALC_VALUE, "trace"

    if normalized.startswith("<"):
        return raw_value, BELOW_LIMIT_CALC_VALUE, "less_than"

    decimal_text = normalized.replace(" ", "")
    if "," in decimal_text and "." not in decimal_text:
        decimal_text = decimal_text.replace(",", ".")

    try:
        return raw_value, Decimal(decimal_text), None
    except InvalidOperation:
        return raw_value, None, "non_numeric"


def read_anses_data() -> tuple[list[SourceFood], list[AnsesNutrientValue], list[str]]:
    """Read ANSES source foods and nutrient values from the workbook."""
    if not ANSES_FILE.exists():
        print(f"File not found: {ANSES_FILE}", file=sys.stderr)
        raise SystemExit(1)

    workbook = load_workbook(ANSES_FILE, read_only=True, data_only=True)
    worksheet = workbook[FOOD_COMPOSITION_SHEET_NAME]
    columns = get_columns(worksheet)
    validate_columns(columns, REQUIRED_FOOD_COLUMNS)

    nutrient_columns, skipped_columns = read_nutrient_column_map(workbook)

    foods: list[SourceFood] = []
    values: list[AnsesNutrientValue] = []

    for row_values in worksheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(columns, row_values))
        food_code = clean_text(row.get(FOOD_CODE_COLUMN))
        food_name_en = clean_text(row.get(FOOD_NAME_EN_COLUMN), dash_is_missing=True)
        if food_code is None:
            continue

        category_en = clean_text(row.get(CATEGORY_EN_COLUMN), dash_is_missing=True)
        foods.append(
            SourceFood(
                source_food_code=food_code,
                food_name_original=food_name_en,
                food_name_en=food_name_en,
                food_name_ro=None,
                category_original=category_en,
                category_en=category_en,
                category_ro=None,
                basis="per_100g",
                notes=build_food_notes(row),
            )
        )

        for nutrient_column in nutrient_columns:
            parsed_value = parse_nutrient_value(row.get(nutrient_column.header))
            if parsed_value is None:
                continue

            raw_value, numeric_value, qualifier = parsed_value
            values.append(
                AnsesNutrientValue(
                    source_food_code=food_code,
                    source_nutrient_code=nutrient_column.source_nutrient_code,
                    raw_value=raw_value,
                    value=numeric_value,
                    value_qualifier=qualifier,
                    unit=nutrient_column.unit,
                    basis="per_100g",
                )
            )

    return foods, values, skipped_columns


def fetch_anses_data_source_id(cursor) -> int:
    cursor.execute(
        "SELECT id FROM data_sources WHERE code = %s",
        (ANSES_DATA_SOURCE_CODE,),
    )
    row = cursor.fetchone()
    if row is None:
        raise RuntimeError(
            "Missing ANSES_CIQUAL row in data_sources. Run migrations through 005 first."
        )
    return int(row["id"])


def fetch_source_food_ids(cursor, data_source_id: int) -> dict[str, int]:
    cursor.execute(
        """
        SELECT id, source_food_code
        FROM source_foods
        WHERE data_source_id = %s
        """,
        (data_source_id,),
    )
    return {row["source_food_code"]: row["id"] for row in cursor.fetchall()}


def fetch_source_nutrient_ids(cursor, data_source_id: int) -> dict[str, int]:
    cursor.execute(
        """
        SELECT id, source_nutrient_code
        FROM source_nutrients
        WHERE data_source_id = %s
        """,
        (data_source_id,),
    )
    return {row["source_nutrient_code"]: row["id"] for row in cursor.fetchall()}


def fetch_reference_id(cursor, data_source_id: int) -> int:
    cursor.execute(
        """
        SELECT id
        FROM source_references
        WHERE data_source_id = %s
          AND source_code = %s
        """,
        (data_source_id, ANSES_REFERENCE_CODE),
    )
    row = cursor.fetchone()
    if row is None:
        raise RuntimeError("Missing ANSES workbook reference row.")
    return int(row["id"])


def validate_import_dependencies(
    values: list[AnsesNutrientValue],
    nutrient_ids: dict[str, int],
) -> None:
    """Fail before writing when ANSES nutrient definitions are missing."""
    missing_nutrient_codes = sorted(
        {value.source_nutrient_code for value in values}
        - set(nutrient_ids)
    )

    if missing_nutrient_codes:
        raise RuntimeError(
            "Missing ANSES source_nutrients rows. First missing codes: "
            + ", ".join(missing_nutrient_codes[:20])
            + ". Run import_source_nutrients.py first."
        )


def iter_chunks(items: list, size: int):
    for index in range(0, len(items), size):
        yield items[index : index + size]


def upsert_reference(cursor, data_source_id: int) -> None:
    """Insert or update the workbook-level ANSES reference row."""
    cursor.execute(
        """
        INSERT INTO source_references (
          data_source_id,
          source_code,
          reference_text
        ) VALUES (
          %s, %s, %s
        ) AS new
        ON DUPLICATE KEY UPDATE
          reference_text = new.reference_text
        """,
        (data_source_id, ANSES_REFERENCE_CODE, ANSES_REFERENCE_TEXT),
    )


def upsert_source_foods(
    cursor,
    data_source_id: int,
    foods: list[SourceFood],
) -> None:
    """Insert or update ANSES source food definitions."""
    sql = """
        INSERT INTO source_foods (
          data_source_id,
          source_food_code,
          food_name_original,
          food_name_en,
          food_name_ro,
          category_original,
          category_en,
          category_ro,
          basis,
          notes
        ) VALUES (
          %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        ) AS new
        ON DUPLICATE KEY UPDATE
          food_name_original = new.food_name_original,
          food_name_en = new.food_name_en,
          food_name_ro = new.food_name_ro,
          category_original = new.category_original,
          category_en = new.category_en,
          category_ro = new.category_ro,
          basis = new.basis,
          notes = new.notes
    """

    rows = [
        (
            data_source_id,
            food.source_food_code,
            food.food_name_original,
            food.food_name_en,
            food.food_name_ro,
            food.category_original,
            food.category_en,
            food.category_ro,
            food.basis,
            food.notes,
        )
        for food in foods
    ]

    for chunk in iter_chunks(rows, INSERT_CHUNK_SIZE):
        cursor.executemany(sql, chunk)


def delete_existing_anses_values(cursor, data_source_id: int) -> None:
    """Remove existing ANSES value rows so the import is repeatable."""
    cursor.execute(
        """
        DELETE v
        FROM source_food_nutrient_values v
        JOIN source_foods f ON f.id = v.source_food_id
        WHERE f.data_source_id = %s
        """,
        (data_source_id,),
    )


def insert_values(
    cursor,
    values: list[AnsesNutrientValue],
    food_ids: dict[str, int],
    nutrient_ids: dict[str, int],
    reference_id: int,
) -> None:
    """Insert ANSES nutrient value rows."""
    sql = """
        INSERT INTO source_food_nutrient_values (
          source_food_id,
          source_nutrient_id,
          raw_value,
          value,
          value_qualifier,
          unit,
          basis,
          reference_id
        ) VALUES (
          %s, %s, %s, %s, %s, %s, %s, %s
        )
    """
    rows = [
        (
            food_ids[value.source_food_code],
            nutrient_ids[value.source_nutrient_code],
            value.raw_value,
            value.value,
            value.value_qualifier,
            value.unit,
            value.basis,
            reference_id,
        )
        for value in values
    ]

    for chunk in iter_chunks(rows, INSERT_CHUNK_SIZE):
        cursor.executemany(sql, chunk)


def print_summary(
    foods: list[SourceFood],
    values: list[AnsesNutrientValue],
    skipped_columns: list[str],
) -> None:
    """Print import summary for review."""
    category_counts = Counter(food.category_en or "(missing)" for food in foods)
    qualifier_counts = Counter(value.value_qualifier or "(blank)" for value in values)
    unit_counts = Counter(value.unit or "(missing)" for value in values)

    print("ANSES/Ciqual import summary")
    print("=" * 40)
    print(f"Foods to import: {len(foods)}")
    print(f"Nutrient value rows to import: {len(values)}")
    print(f"Unique nutrients in values: {len({value.source_nutrient_code for value in values})}")
    print(f"Skipped nutrient columns: {len(skipped_columns)}")

    if skipped_columns:
        print()
        print("Skipped columns")
        print("=" * 40)
        for column in skipped_columns:
            print(f"- {clean_text(column) or column}")

    print()
    print("Top categories")
    print("=" * 40)
    for category, count in category_counts.most_common(10):
        print(f"{category}: {count}")

    print()
    print("Value qualifier counts")
    print("=" * 40)
    for qualifier, count in qualifier_counts.most_common():
        print(f"{qualifier}: {count}")

    print()
    print("Top units")
    print("=" * 40)
    for unit, count in unit_counts.most_common(10):
        print(f"{unit}: {count}")

    print()
    print("First 5 foods")
    print("=" * 40)
    for food in foods[:5]:
        print(
            f"- {food.source_food_code} | "
            f"{food.food_name_en or '-'} | "
            f"{food.category_en or '-'} | "
            f"{food.basis or '-'}"
        )

    print()
    print("First 5 values")
    print("=" * 40)
    for value in values[:5]:
        print(
            f"- food={value.source_food_code}, "
            f"nutrient={value.source_nutrient_code}, "
            f"raw={value.raw_value}, "
            f"value={value.value}, "
            f"qualifier={value.value_qualifier or '-'}, "
            f"unit={value.unit or '-'}"
        )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Import ANSES/Ciqual foods and nutrient values."
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Read files and validate database dependencies without writing.",
    )
    args = parser.parse_args()

    foods, values, skipped_columns = read_anses_data()
    print_summary(foods, values, skipped_columns)

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            data_source_id = fetch_anses_data_source_id(cursor)
            nutrient_ids = fetch_source_nutrient_ids(cursor, data_source_id)
            validate_import_dependencies(values, nutrient_ids)

            if args.dry_run:
                print()
                print("Dry run only. No database changes were written.")
                return 0

            upsert_reference(cursor, data_source_id)
            upsert_source_foods(cursor, data_source_id, foods)
            food_ids = fetch_source_food_ids(cursor, data_source_id)
            reference_id = fetch_reference_id(cursor, data_source_id)

            delete_existing_anses_values(cursor, data_source_id)
            insert_values(cursor, values, food_ids, nutrient_ids, reference_id)
            connection.commit()

            print()
            print(f"Imported {len(foods)} ANSES source food definitions.")
            print(f"Imported {len(values)} ANSES nutrient value rows.")
            return 0
    except Exception as exc:
        connection.rollback()
        print(f"ERROR: ANSES import failed: {exc}", file=sys.stderr)
        return 1
    finally:
        connection.close()


if __name__ == "__main__":
    raise SystemExit(main())
