#!/usr/bin/env python3
"""Compare ANSES/Ciqual and NEVO nutrient definitions.

This script does not modify the database. It reads the source files and prints
code/name/unit overlaps so we can design the external-source schema with less
guesswork.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
import re

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[3]

ANSES_FILE = (
    PROJECT_ROOT
    / "temp"
    / "EuropeNutrientsDBs"
    / "anses"
    / "Table Ciqual 2025_ENG_2025_11_03.xlsx"
)
NEVO_NUTRIENTS_FILE = (
    PROJECT_ROOT
    / "temp"
    / "EuropeNutrientsDBs"
    / "nevo"
    / "NEVO2025_v9.0_Nutrienten_Nutrients.csv"
)

ANSES_SHEET_NAME = "INFOODS codes"
ANSES_INFOODS_TAG_COLUMN = "INFDSTAG"
ANSES_SOURCE_CODE_COLUMN = "ORIGCPCD"
ANSES_NAME_COLUMN = "const_nom_eng"

NEVO_CODE_COLUMN = "Nutrient-code"
NEVO_NAME_COLUMN = "Component"
NEVO_UNIT_COLUMN = "Eenheid/Unit"
NEVO_GROUP_COLUMN = "Component group"


@dataclass
class NutrientDefinition:
    """Source nutrient definition used for cross-source comparison."""

    source: str
    source_code: str
    name: str
    unit: str
    infoods_tag: str = ""
    groups: set[str] = field(default_factory=set)


def normalize_unit(value: str) -> str:
    """Normalize unit spelling for comparison."""
    cleaned = value.strip().lower()
    cleaned = cleaned.replace("\u00c2\u00b5", "u")
    cleaned = cleaned.replace("\u00b5", "u")
    cleaned = cleaned.replace("\u03bc", "u")
    return cleaned


def normalize_name(value: str) -> str:
    """Normalize nutrient names without trying to fully standardize meaning."""
    cleaned = value.strip().lower()
    cleaned = cleaned.replace("\n", " ")
    cleaned = cleaned.replace("'", "")
    cleaned = cleaned.replace("-", " ")
    cleaned = re.sub(r"[^a-z0-9]+", " ", cleaned)
    cleaned = re.sub(r"\b(total|available|dietary)\b", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip()


def parse_anses_name_and_unit(value: str) -> tuple[str, str]:
    """Split an ANSES label like 'Water (g/100g)' into name and unit."""
    match = re.search(r"\(([^()]*)\)\s*$", value)

    if not match:
        return value.strip(), ""

    raw_unit = match.group(1)
    name = value[: match.start()].strip()
    unit = raw_unit.split("/", 1)[0].strip()
    return name, normalize_unit(unit)


def read_anses_nutrients() -> list[NutrientDefinition]:
    """Read ANSES nutrient definitions from the INFOODS sheet."""
    workbook = load_workbook(ANSES_FILE, read_only=True, data_only=True)
    worksheet = workbook[ANSES_SHEET_NAME]
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))

    infoods_index = headers.index(ANSES_INFOODS_TAG_COLUMN)
    source_code_index = headers.index(ANSES_SOURCE_CODE_COLUMN)
    name_index = headers.index(ANSES_NAME_COLUMN)

    nutrients: list[NutrientDefinition] = []

    for row in rows:
        full_name = str(row[name_index]).strip()
        name, unit = parse_anses_name_and_unit(full_name)
        infoods_tag = "" if row[infoods_index] is None else str(row[infoods_index]).strip()

        nutrients.append(
            NutrientDefinition(
                source="ANSES",
                source_code=str(row[source_code_index]).strip(),
                name=name,
                unit=unit,
                infoods_tag=infoods_tag,
            )
        )

    return nutrients


def read_nevo_nutrients() -> list[NutrientDefinition]:
    """Read unique NEVO nutrient definitions from the nutrient dictionary."""
    by_code: dict[str, NutrientDefinition] = {}

    with NEVO_NUTRIENTS_FILE.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file, delimiter="|")

        for row in reader:
            code = row[NEVO_CODE_COLUMN].strip()
            name = row[NEVO_NAME_COLUMN].strip()
            unit = normalize_unit(row[NEVO_UNIT_COLUMN])
            group = row[NEVO_GROUP_COLUMN].strip()

            if code not in by_code:
                by_code[code] = NutrientDefinition(
                    source="NEVO",
                    source_code=code,
                    name=name,
                    unit=unit,
                    groups={group} if group else set(),
                )
            elif group:
                by_code[code].groups.add(group)

    return list(by_code.values())


def similarity(left: NutrientDefinition, right: NutrientDefinition) -> float:
    """Score rough semantic similarity from name and unit."""
    if has_conflicting_identifier(left.name.lower(), right.name.lower()):
        return 0

    left_name = normalize_name(left.name)
    right_name = normalize_name(right.name)
    name_score = SequenceMatcher(None, left_name, right_name).ratio()

    if left.unit and right.unit and left.unit == right.unit:
        return min(name_score + 0.15, 1)

    if left.unit and right.unit and left.unit != right.unit:
        return name_score - 0.15

    return name_score


def has_conflicting_identifier(left_name: str, right_name: str) -> bool:
    """Reject likely matches with conflicting vitamin or fatty-acid markers."""
    left_vitamins = set(re.findall(r"\bvit(?:amin)?\s*([a-z]\d*)\b", left_name))
    right_vitamins = set(re.findall(r"\bvit(?:amin)?\s*([a-z]\d*)\b", right_name))

    if left_vitamins and right_vitamins and left_vitamins.isdisjoint(right_vitamins):
        return True

    left_fatty_acids = set(re.findall(r"\b(?:fa|c)?\s*(\d{1,2}:\d)\b", left_name))
    right_fatty_acids = set(re.findall(r"\b(?:fa|c)?\s*(\d{1,2}:\d)\b", right_name))

    if (
        left_fatty_acids
        and right_fatty_acids
        and left_fatty_acids.isdisjoint(right_fatty_acids)
    ):
        return True

    return False


def find_direct_code_matches(
    anses_nutrients: list[NutrientDefinition],
    nevo_nutrients: list[NutrientDefinition],
) -> list[tuple[NutrientDefinition, NutrientDefinition]]:
    """Match ANSES INFOODS tags directly to NEVO nutrient codes."""
    nevo_by_code = {nutrient.source_code: nutrient for nutrient in nevo_nutrients}
    matches = []

    for anses_nutrient in anses_nutrients:
        if anses_nutrient.infoods_tag in nevo_by_code:
            matches.append((anses_nutrient, nevo_by_code[anses_nutrient.infoods_tag]))

    return matches


def find_exact_name_unit_matches(
    anses_nutrients: list[NutrientDefinition],
    nevo_nutrients: list[NutrientDefinition],
) -> list[tuple[NutrientDefinition, NutrientDefinition]]:
    """Match normalized names and units exactly."""
    nevo_by_name_unit = {
        (normalize_name(nutrient.name), nutrient.unit): nutrient
        for nutrient in nevo_nutrients
    }
    matches = []

    for anses_nutrient in anses_nutrients:
        key = (normalize_name(anses_nutrient.name), anses_nutrient.unit)
        nevo_nutrient = nevo_by_name_unit.get(key)

        if nevo_nutrient:
            matches.append((anses_nutrient, nevo_nutrient))

    return matches


def find_likely_matches(
    anses_nutrients: list[NutrientDefinition],
    nevo_nutrients: list[NutrientDefinition],
    threshold: float = 0.82,
) -> list[tuple[float, NutrientDefinition, NutrientDefinition]]:
    """Find likely name/unit matches even when source codes differ."""
    direct_matches = {
        anses.source_code
        for anses, _nevo in find_direct_code_matches(anses_nutrients, nevo_nutrients)
    }
    exact_matches = {
        anses.source_code
        for anses, _nevo in find_exact_name_unit_matches(anses_nutrients, nevo_nutrients)
    }
    matches = []

    for anses_nutrient in anses_nutrients:
        if anses_nutrient.source_code in direct_matches | exact_matches:
            continue

        scored_candidates = [
            (similarity(anses_nutrient, nevo_nutrient), nevo_nutrient)
            for nevo_nutrient in nevo_nutrients
        ]
        score, best_nevo = max(scored_candidates, key=lambda item: item[0])

        if score >= threshold:
            matches.append((score, anses_nutrient, best_nevo))

    matches.sort(key=lambda item: item[0], reverse=True)
    return matches


def print_nutrient_pair(
    anses_nutrient: NutrientDefinition,
    nevo_nutrient: NutrientDefinition,
    prefix: str = "-",
) -> None:
    """Print one ANSES/NEVO nutrient pair."""
    anses_tag = anses_nutrient.infoods_tag or "(no INFDSTAG)"
    print(
        f"{prefix} ANSES {anses_nutrient.source_code} / {anses_tag} | "
        f"{anses_nutrient.name} ({anses_nutrient.unit})"
    )
    print(
        f"  NEVO  {nevo_nutrient.source_code} | "
        f"{nevo_nutrient.name} ({nevo_nutrient.unit})"
    )


def main() -> int:
    """Run the nutrient comparison."""
    if not ANSES_FILE.exists():
        print(f"File not found: {ANSES_FILE}")
        return 1

    if not NEVO_NUTRIENTS_FILE.exists():
        print(f"File not found: {NEVO_NUTRIENTS_FILE}")
        return 1

    anses_nutrients = read_anses_nutrients()
    nevo_nutrients = read_nevo_nutrients()

    direct_code_matches = find_direct_code_matches(anses_nutrients, nevo_nutrients)
    exact_name_unit_matches = find_exact_name_unit_matches(anses_nutrients, nevo_nutrients)
    likely_matches = find_likely_matches(anses_nutrients, nevo_nutrients)

    direct_anses_codes = {anses.source_code for anses, _nevo in direct_code_matches}
    exact_anses_codes = {anses.source_code for anses, _nevo in exact_name_unit_matches}
    likely_anses_codes = {anses.source_code for _score, anses, _nevo in likely_matches}
    covered_anses_codes = direct_anses_codes | exact_anses_codes | likely_anses_codes

    print("ANSES vs NEVO nutrient comparison")
    print("=" * 40)
    print(f"ANSES nutrient definitions: {len(anses_nutrients)}")
    print(f"NEVO unique nutrient definitions: {len(nevo_nutrients)}")
    print(f"Direct ANSES INFDSTAG -> NEVO code matches: {len(direct_code_matches)}")
    print(f"Exact normalized name/unit matches: {len(exact_name_unit_matches)}")
    print(f"Likely name/unit matches with different codes: {len(likely_matches)}")
    print(f"ANSES nutrients covered by these checks: {len(covered_anses_codes)}")
    print()

    print("Direct code matches")
    print("=" * 40)
    for anses_nutrient, nevo_nutrient in direct_code_matches[:30]:
        print_nutrient_pair(anses_nutrient, nevo_nutrient)

    print()
    print("Exact name/unit matches with different codes")
    print("=" * 40)
    exact_different_code_matches = [
        (anses_nutrient, nevo_nutrient)
        for anses_nutrient, nevo_nutrient in exact_name_unit_matches
        if anses_nutrient.infoods_tag != nevo_nutrient.source_code
    ]
    for anses_nutrient, nevo_nutrient in exact_different_code_matches[:30]:
        print_nutrient_pair(anses_nutrient, nevo_nutrient)

    print()
    print("Likely matches with different codes")
    print("=" * 40)
    for score, anses_nutrient, nevo_nutrient in likely_matches[:30]:
        print(f"- score={score:.2f}")
        print_nutrient_pair(anses_nutrient, nevo_nutrient, prefix="  ")

    print()
    print("ANSES nutrients not covered by direct/exact/likely checks")
    print("=" * 40)
    uncovered = [
        nutrient
        for nutrient in anses_nutrients
        if nutrient.source_code not in covered_anses_codes
    ]

    for nutrient in uncovered[:40]:
        tag = nutrient.infoods_tag or "(no INFDSTAG)"
        print(f"- {nutrient.source_code} / {tag} | {nutrient.name} ({nutrient.unit})")

    print(f"Total uncovered ANSES nutrients: {len(uncovered)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
