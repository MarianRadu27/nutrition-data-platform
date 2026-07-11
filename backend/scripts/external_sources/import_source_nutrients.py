#!/usr/bin/env python3
"""Import external-source nutrient definitions into source_nutrients.

This imports nutrient definitions only, not food rows and not nutrient values.
Use --dry-run first to preview counts without writing to the database.
"""

from __future__ import annotations

import argparse
import csv
import os
from dataclasses import dataclass
from pathlib import Path
import re
import sys

import pymysql
from dotenv import load_dotenv
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[3]
BACKEND_DIR = PROJECT_ROOT / "backend"

ANSES_FILE = (
    PROJECT_ROOT
    / "temp"
    / "EuropeNutrientsDBs"
    / "anses"
    / "Table Ciqual 2025_ENG_2025_11_03.xlsx"
)
NEVO_NUTRIENTS_FILE = (
    PROJECT_ROOT
    / "temp"
    / "EuropeNutrientsDBs"
    / "nevo"
    / "NEVO2025_v9.0_Nutrienten_Nutrients.csv"
)

ANSES_DATA_SOURCE_CODE = "ANSES_CIQUAL"
NEVO_DATA_SOURCE_CODE = "NEVO"

ANSES_SHEET_NAME = "INFOODS codes"
ANSES_INFOODS_TAG_COLUMN = "INFDSTAG"
ANSES_SOURCE_CODE_COLUMN = "ORIGCPCD"
ANSES_NAME_COLUMN = "const_nom_eng"

NEVO_CODE_COLUMN = "Nutrient-code"
NEVO_NAME_COLUMN = "Component"
NEVO_UNIT_COLUMN = "Eenheid/Unit"
NEVO_GROUP_COLUMN = "Component group"

ANSES_CANONICAL_MAPPING = {
    "328": "energy_kcal",
    "25000": "protein_g",
    "31000": "carbohydrate_g",
    "40000": "fat_g",
    "34100": "fiber_g",
    "32000": "sugar_g",
    "10004": "salt_g",
    "10110": "sodium_mg",
    "400": "water_g",
}

NEVO_CANONICAL_MAPPING = {
    "ENERCC": "energy_kcal",
    "PROT": "protein_g",
    "CHO": "carbohydrate_g",
    "FAT": "fat_g",
    "FIBT": "fiber_g",
    "SUGAR": "sugar_g",
    "NA": "sodium_mg",
    "WATER": "water_g",
}


@dataclass(frozen=True)
class SourceNutrient:
    """One nutrient definition from one source."""

    data_source_code: str
    source_nutrient_code: str
    source_nutrient_name: str
    source_standard_tag: str | None
    unit: str | None
    component_group: str | None
    canonical_code: str | None


def env_str(name: str, default: str) -> str:
    value = os.getenv(name)
    if value is None:
        return default
    stripped = value.strip()
    return stripped if stripped else default


def env_int(name: str, default: int) -> int:
    raw = os.getenv(name)
    if raw is None:
        return default
    try:
        return int(raw.strip())
    except ValueError:
        return default


def get_connection() -> pymysql.connections.Connection:
    """Open a MySQL connection using the backend .env settings."""
    load_dotenv(BACKEND_DIR / ".env")
    return pymysql.connect(
        host=env_str("DB_HOST", "127.0.0.1"),
        port=env_int("DB_PORT", 3307),
        user=env_str("DB_USER", "nutrition"),
        password=env_str("DB_PASSWORD", "nutritionpass"),
        database=env_str("DB_NAME", "nutrition"),
        charset="utf8mb4",
        autocommit=False,
        cursorclass=pymysql.cursors.DictCursor,
    )


def normalize_unit(value: str | None) -> str | None:
    """Normalize unit spelling while keeping storage simple."""
    if value is None:
        return None

    cleaned = value.strip()
    if not cleaned:
        return None

    cleaned = cleaned.replace("\u00c2\u00b5", "u")
    cleaned = cleaned.replace("\u00b5", "u")
    cleaned = cleaned.replace("\u03bc", "u")
    cleaned = cleaned.replace("�", "u")
    return cleaned.lower()


def parse_anses_unit(source_name: str) -> str | None:
    """Extract the reported unit from an ANSES label such as '(g/100g)'."""
    match = re.search(r"\(([^()]*)\)\s*$", source_name)
    if not match:
        return None

    raw_unit = match.group(1).split("/", 1)[0]
    return normalize_unit(raw_unit)


def read_anses_source_nutrients() -> list[SourceNutrient]:
    """Read ANSES nutrient definitions from the INFOODS codes sheet."""
    workbook = load_workbook(ANSES_FILE, read_only=True, data_only=True)
    worksheet = workbook[ANSES_SHEET_NAME]
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))

    tag_index = headers.index(ANSES_INFOODS_TAG_COLUMN)
    code_index = headers.index(ANSES_SOURCE_CODE_COLUMN)
    name_index = headers.index(ANSES_NAME_COLUMN)

    nutrients: list[SourceNutrient] = []
    for row in rows:
        source_code = str(row[code_index]).strip()
        source_name = str(row[name_index]).strip()
        raw_tag = row[tag_index]
        source_standard_tag = None if raw_tag is None else str(raw_tag).strip() or None

        nutrients.append(
            SourceNutrient(
                data_source_code=ANSES_DATA_SOURCE_CODE,
                source_nutrient_code=source_code,
                source_nutrient_name=source_name,
                source_standard_tag=source_standard_tag,
                unit=parse_anses_unit(source_name),
                component_group=None,
                canonical_code=ANSES_CANONICAL_MAPPING.get(source_code),
            )
        )

    return nutrients


def read_nevo_source_nutrients() -> list[SourceNutrient]:
    """Read unique NEVO nutrient definitions from the nutrient dictionary."""
    nutrients_by_code: dict[str, SourceNutrient] = {}
    groups_by_code: dict[str, set[str]] = {}

    with NEVO_NUTRIENTS_FILE.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file, delimiter="|")

        for row in reader:
            source_code = row[NEVO_CODE_COLUMN].strip()
            source_name = row[NEVO_NAME_COLUMN].strip()
            unit = normalize_unit(row[NEVO_UNIT_COLUMN])
            group = row[NEVO_GROUP_COLUMN].strip()

            groups_by_code.setdefault(source_code, set())
            if group:
                groups_by_code[source_code].add(group)

            if source_code not in nutrients_by_code:
                nutrients_by_code[source_code] = SourceNutrient(
                    data_source_code=NEVO_DATA_SOURCE_CODE,
                    source_nutrient_code=source_code,
                    source_nutrient_name=source_name,
                    source_standard_tag=None,
                    unit=unit,
                    component_group=None,
                    canonical_code=NEVO_CANONICAL_MAPPING.get(source_code),
                )

    nutrients: list[SourceNutrient] = []
    for source_code, nutrient in nutrients_by_code.items():
        component_group = "; ".join(sorted(groups_by_code[source_code])) or None
        nutrients.append(
            SourceNutrient(
                data_source_code=nutrient.data_source_code,
                source_nutrient_code=nutrient.source_nutrient_code,
                source_nutrient_name=nutrient.source_nutrient_name,
                source_standard_tag=nutrient.source_standard_tag,
                unit=nutrient.unit,
                component_group=component_group,
                canonical_code=nutrient.canonical_code,
            )
        )

    return nutrients


def read_all_source_nutrients() -> list[SourceNutrient]:
    """Read all supported source nutrient definitions."""
    missing_files = [
        path
        for path in (ANSES_FILE, NEVO_NUTRIENTS_FILE)
        if not path.exists()
    ]
    if missing_files:
        for path in missing_files:
            print(f"File not found: {path}", file=sys.stderr)
        raise SystemExit(1)

    return [
        *read_nevo_source_nutrients(),
        *read_anses_source_nutrients(),
    ]


def fetch_data_source_ids(cursor) -> dict[str, int]:
    cursor.execute("SELECT id, code FROM data_sources")
    return {row["code"]: row["id"] for row in cursor.fetchall()}


def fetch_canonical_nutrient_ids(cursor) -> dict[str, int]:
    cursor.execute("SELECT id, canonical_code FROM canonical_nutrients")
    return {row["canonical_code"]: row["id"] for row in cursor.fetchall()}


def validate_required_seed_data(
    nutrients: list[SourceNutrient],
    data_source_ids: dict[str, int],
    canonical_ids: dict[str, int],
) -> None:
    """Fail early when required seed migrations were not applied."""
    missing_sources = sorted(
        {nutrient.data_source_code for nutrient in nutrients}
        - set(data_source_ids)
    )
    missing_canonical = sorted(
        {
            nutrient.canonical_code
            for nutrient in nutrients
            if nutrient.canonical_code is not None
        }
        - set(canonical_ids)
    )

    if missing_sources:
        raise RuntimeError(
            "Missing data_sources rows: "
            + ", ".join(missing_sources)
            + ". Run migrations through 005 first."
        )

    if missing_canonical:
        raise RuntimeError(
            "Missing canonical_nutrients rows: "
            + ", ".join(missing_canonical)
            + ". Run migrations through 004 first."
        )


def upsert_source_nutrients(
    cursor,
    nutrients: list[SourceNutrient],
    data_source_ids: dict[str, int],
    canonical_ids: dict[str, int],
) -> None:
    """Insert or update source nutrient definitions."""
    sql = """
        INSERT INTO source_nutrients (
          data_source_id,
          source_nutrient_code,
          source_nutrient_name,
          source_standard_tag,
          unit,
          component_group,
          canonical_nutrient_id
        ) VALUES (
          %s, %s, %s, %s, %s, %s, %s
        ) AS new
        ON DUPLICATE KEY UPDATE
          source_nutrient_name = new.source_nutrient_name,
          source_standard_tag = new.source_standard_tag,
          unit = new.unit,
          component_group = new.component_group,
          canonical_nutrient_id = new.canonical_nutrient_id
    """

    rows = [
        (
            data_source_ids[nutrient.data_source_code],
            nutrient.source_nutrient_code,
            nutrient.source_nutrient_name,
            nutrient.source_standard_tag,
            nutrient.unit,
            nutrient.component_group,
            canonical_ids.get(nutrient.canonical_code or ""),
        )
        for nutrient in nutrients
    ]
    cursor.executemany(sql, rows)


def print_summary(nutrients: list[SourceNutrient]) -> None:
    """Print import summary for review."""
    by_source: dict[str, list[SourceNutrient]] = {}
    for nutrient in nutrients:
        by_source.setdefault(nutrient.data_source_code, []).append(nutrient)

    print("Source nutrient import summary")
    print("=" * 36)
    for source_code, source_nutrients in sorted(by_source.items()):
        mapped = sum(1 for nutrient in source_nutrients if nutrient.canonical_code)
        print(f"{source_code}: {len(source_nutrients)} nutrients, {mapped} canonical mappings")

    print()
    print("Canonical mappings")
    print("=" * 36)
    for nutrient in nutrients:
        if nutrient.canonical_code:
            tag = nutrient.source_standard_tag or "-"
            print(
                f"- {nutrient.data_source_code} | "
                f"{nutrient.source_nutrient_code} | "
                f"{tag} | "
                f"{nutrient.source_nutrient_name} | "
                f"{nutrient.unit or '-'} -> "
                f"{nutrient.canonical_code}"
            )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Import NEVO and ANSES nutrient definitions into source_nutrients."
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Read files and validate seed data without writing to the database.",
    )
    args = parser.parse_args()

    nutrients = read_all_source_nutrients()
    print_summary(nutrients)

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            data_source_ids = fetch_data_source_ids(cursor)
            canonical_ids = fetch_canonical_nutrient_ids(cursor)
            validate_required_seed_data(nutrients, data_source_ids, canonical_ids)

            if args.dry_run:
                print()
                print("Dry run only. No database changes were written.")
                return 0

            upsert_source_nutrients(cursor, nutrients, data_source_ids, canonical_ids)
            connection.commit()
            print()
            print(f"Imported {len(nutrients)} source nutrient definitions.")
            return 0
    except Exception as exc:
        connection.rollback()
        print(f"ERROR: source nutrient import failed: {exc}", file=sys.stderr)
        return 1
    finally:
        connection.close()


if __name__ == "__main__":
    raise SystemExit(main())
